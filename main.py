import pyshark
from collections import Counter
import matplotlib.pyplot as plt
import pandas as pd
from matplotlib.ticker import MaxNLocator
import numpy as np
from urllib.parse import urlparse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import datetime


def analyze_network_traffic(pcap_file, top_n=100):
    # Configurações iniciais
    plt.style.use('ggplot')
    pd.set_option('display.max_rows', None)

    # Contadores
    full_domains = Counter()
    base_domains = Counter()
    ip_users = Counter()  # Contador para endereços IP únicos (usuários)

    print("Processando arquivo pcap, isso pode demorar...")

    try:
        cap = pyshark.FileCapture(
            pcap_file,
            display_filter='http or dns or (tcp and tcp.flags.syn==1 and tcp.flags.ack==0)',
            only_summaries=False
        )

        for i, packet in enumerate(cap):
            try:
                # Contagem de usuários (conexões TCP SYN)
                if 'tcp' in packet and hasattr(packet.tcp, 'flags_syn') and packet.tcp.flags_syn == '1':
                    ip_src = packet.ip.src
                    ip_users[ip_src] += 1

                # Análise HTTP
                elif 'http' in packet and hasattr(packet.http, 'host'):
                    host = packet.http.host
                    full_domains[host] += 1

                    parsed = urlparse(f'http://{host}')
                    domain = parsed.netloc.split(':')[0]
                    if domain.count('.') > 1:
                        domain = '.'.join(domain.split('.')[-2:])
                    base_domains[domain] += 1

                # Análise DNS
                elif 'dns' in packet and hasattr(packet.dns, 'qry_name'):
                    domain = packet.dns.qry_name
                    if not domain.endswith('.arpa'):
                        full_domains[domain] += 1
                        if domain.count('.') > 1:
                            base_domain = '.'.join(domain.split('.')[-2:])
                            base_domains[base_domain] += 1

            except Exception as e:
                continue

            if i % 10000 == 0:
                print(f"Pacotes processados: {i}")

    except Exception as e:
        print(f"Erro ao processar arquivo: {e}")
        return None
    finally:
        cap.close()

    print("\nProcessamento concluído!\n")

    # Resultados
    top_full = full_domains.most_common(top_n)
    top_base = base_domains.most_common(top_n)
    total_users = len(ip_users)  # Total de IPs únicos

    # Criar DataFrames
    df_full = pd.DataFrame(top_full, columns=['Domínio Completo', 'Acessos'])
    df_base = pd.DataFrame(top_base, columns=['Domínio Principal', 'Acessos'])
    df_users = pd.DataFrame({'Métrica': ['Usuários Únicos'], 'Valor': [total_users]})

    # Gerar relatório Excel
    generate_excel_report(df_full, df_base, df_users, ip_users)

    # Visualização
    plot_top_sites(df_base, title='Top 100 Domínios Principais Mais Acessados')
    plot_top_sites(df_full, title='Top 100 URLs Completas Mais Acessadas')

    return df_full, df_base, df_users


def generate_excel_report(df_full, df_base, df_users, ip_users):
    """Gera um relatório completo em Excel"""
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f'relatorio_rede_{timestamp}.xlsx'

    wb = Workbook()

    # Planilha de domínios completos
    ws_full = wb.active
    ws_full.title = "Domínios Completos"
    for r in dataframe_to_rows(df_full, index=False, header=True):
        ws_full.append(r)

    # Planilha de domínios principais
    ws_base = wb.create_sheet("Domínios Principais")
    for r in dataframe_to_rows(df_base, index=False, header=True):
        ws_base.append(r)

    # Planilha de usuários
    ws_users = wb.create_sheet("Usuários")
    ws_users.append(["Relatório de Usuários Logados"])
    ws_users.append(["Data/Hora", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws_users.append([])

    for r in dataframe_to_rows(df_users, index=False, header=True):
        ws_users.append(r)

    # Lista de IPs únicos
    ws_users.append([])
    ws_users.append(["Endereços IP Únicos Detectados"])
    for ip in ip_users:
        ws_users.append([ip])

    # Formatação
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

    # Salvar
    wb.save(filename)
    print(f"\nRelatório Excel gerado: {filename}")


def plot_top_sites(df, title):
    """Gera gráficos de barras horizontais"""
    plt.figure(figsize=(14, 20))
    df = df.sort_values('Acessos', ascending=True)
    colors = plt.cm.viridis(np.linspace(0.2, 0.8, len(df)))

    bars = plt.barh(
        df.iloc[:, 0],
        df['Acessos'],
        color=colors,
        edgecolor='black',
        linewidth=0.5
    )

    for bar in bars:
        width = bar.get_width()
        plt.text(
            width + max(df['Acessos']) * 0.01,
            bar.get_y() + bar.get_height() / 2,
            f'{int(width):,}',
            va='center',
            ha='left',
            fontsize=9
        )

    plt.title(title, fontsize=14, pad=20)
    plt.xlabel('Número de Acessos', fontsize=12)
    plt.ylabel('Domínios', fontsize=12)
    plt.gca().xaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{int(x):,}'))
    plt.gca().xaxis.set_major_locator(MaxNLocator(integer=True))
    plt.grid(axis='x', alpha=0.4)
    plt.tight_layout()

    filename = title.lower().replace(' ', '_').replace('ç', 'c') + '.png'
    plt.savefig(filename, dpi=300, bbox_inches='tight')
    plt.show()
    print(f"Gráfico salvo como: {filename}")


# Uso
df_full, df_base, df_users = analyze_network_traffic('captura.pcapng', top_n=100)

# Mostrar resultados no console
print("\n=== RESUMO ===")
print(f"Total de usuários únicos detectados: {df_users['Valor'][0]}")
print("\nTop 10 Domínios Principais:")
print(df_base.head(10))
print("\nTop 10 URLs Completas:")
print(df_full.head(10))