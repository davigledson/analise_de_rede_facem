import pyshark
from collections import defaultdict
import matplotlib.pyplot as plt
import pandas as pd
from matplotlib.ticker import MaxNLocator
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import datetime
import statistics


def analyze_packet_loss(pcap_file, analysis_interval=5):
    """
    Analisa a perda de pacotes na rede com base em um arquivo pcap/pcapng.

    Args:
        pcap_file (str): Caminho para o arquivo de captura
        analysis_interval (int): Intervalo em segundos para análise temporal

    Returns:
        dict: Dicionário com estatísticas de perda de pacotes
    """
    # Configurações iniciais
    plt.style.use('ggplot')
    pd.set_option('display.max_rows', None)

    print("Processando arquivo pcap para análise de perda de pacotes...")

    # Estruturas para armazenar dados
    packet_stats = {
        'total_packets': 0,
        'tcp_packets': 0,
        'udp_packets': 0,
        'icmp_packets': 0,
        'other_packets': 0,
        'retransmissions': 0,
        'out_of_order': 0,
        'duplicate_acks': 0,
        'tcp_streams': defaultdict(dict),
        'time_intervals': defaultdict(lambda: {
            'packets': 0,
            'retransmissions': 0,
            'out_of_order': 0,
            'bytes': 0
        })
    }

    # Contadores para sequência TCP
    tcp_sequence = defaultdict(lambda: {
        'last_seq': None,
        'last_ack': None,
        'packet_count': 0,
        'retrans_count': 0,
        'out_of_order_count': 0,
        'dup_ack_count': 0
    })

    try:
        cap = pyshark.FileCapture(
            pcap_file,
            display_filter='tcp or udp or icmp',
            only_summaries=False
        )

        start_time = None
        for i, packet in enumerate(cap):
            try:
                # Determinar o timestamp do pacote
                if start_time is None:
                    start_time = float(packet.sniff_timestamp)
                current_time = float(packet.sniff_timestamp)
                elapsed = current_time - start_time
                interval = int(elapsed // analysis_interval)

                # Atualizar estatísticas gerais
                packet_stats['total_packets'] += 1
                packet_stats['time_intervals'][interval]['packets'] += 1

                # Analisar por protocolo
                if 'tcp' in packet:
                    packet_stats['tcp_packets'] += 1
                    analyze_tcp_packet(packet, packet_stats, tcp_sequence, interval)
                elif 'udp' in packet:
                    packet_stats['udp_packets'] += 1
                elif 'icmp' in packet:
                    packet_stats['icmp_packets'] += 1
                else:
                    packet_stats['other_packets'] += 1

                # Contar bytes totais
                if hasattr(packet, 'length'):
                    packet_stats['time_intervals'][interval]['bytes'] += int(packet.length)

                if i % 10000 == 0:
                    print(f"Pacotes processados: {i}")

            except Exception as e:
                continue

    except Exception as e:
        print(f"Erro ao processar arquivo: {e}")
        return None
    finally:
        cap.close()

    print("\nProcessamento concluído!\n")

    # Calcular estatísticas resumidas
    packet_stats = calculate_summary_stats(packet_stats, analysis_interval)

    # Gerar relatórios e gráficos
    generate_packet_loss_reports(packet_stats)

    return packet_stats


def analyze_tcp_packet(packet, packet_stats, tcp_sequence, interval):
    """Analisa um pacote TCP para detectar problemas de rede"""
    tcp = packet.tcp
    stream_key = f"{packet.ip.src}:{tcp.srcport}-{packet.ip.dst}:{tcp.dstport}"

    # Verificar retransmissões (usando flags e números de sequência)
    if hasattr(tcp, 'analysis_retransmission') and tcp.analysis_retransmission == '1':
        packet_stats['retransmissions'] += 1
        packet_stats['time_intervals'][interval]['retransmissions'] += 1
        tcp_sequence[stream_key]['retrans_count'] += 1

    # Verificar pacotes fora de ordem
    if hasattr(tcp, 'analysis_out_of_order') and tcp.analysis_out_of_order == '1':
        packet_stats['out_of_order'] += 1
        packet_stats['time_intervals'][interval]['out_of_order'] += 1
        tcp_sequence[stream_key]['out_of_order_count'] += 1

    # Verificar ACKs duplicados
    if hasattr(tcp, 'analysis_duplicate_ack') and tcp.analysis_duplicate_ack == '1':
        packet_stats['duplicate_acks'] += 1
        tcp_sequence[stream_key]['dup_ack_count'] += 1

    # Atualizar informações de sequência para o fluxo TCP
    if hasattr(tcp, 'seq'):
        current_seq = int(tcp.seq)
        last_seq = tcp_sequence[stream_key]['last_seq']

        if last_seq is not None and current_seq < last_seq:
            packet_stats['out_of_order'] += 1
            packet_stats['time_intervals'][interval]['out_of_order'] += 1
            tcp_sequence[stream_key]['out_of_order_count'] += 1

        tcp_sequence[stream_key]['last_seq'] = current_seq

    if hasattr(tcp, 'ack'):
        tcp_sequence[stream_key]['last_ack'] = int(tcp.ack)

    tcp_sequence[stream_key]['packet_count'] += 1

    # Armazenar informações do fluxo TCP
    if stream_key not in packet_stats['tcp_streams']:
        packet_stats['tcp_streams'][stream_key] = {
            'src': packet.ip.src,
            'src_port': tcp.srcport,
            'dst': packet.ip.dst,
            'dst_port': tcp.dstport,
            'packet_count': 0,
            'retrans_count': 0,
            'out_of_order_count': 0,
            'dup_ack_count': 0
        }

    packet_stats['tcp_streams'][stream_key]['packet_count'] += 1


def calculate_summary_stats(packet_stats, interval_seconds):
    """Calcula estatísticas resumidas a partir dos dados coletados"""
    # Calcular taxas de problemas
    if packet_stats['tcp_packets'] > 0:
        packet_stats['retransmission_rate'] = (packet_stats['retransmissions'] / packet_stats['tcp_packets']) * 100
        packet_stats['out_of_order_rate'] = (packet_stats['out_of_order'] / packet_stats['tcp_packets']) * 100
        packet_stats['duplicate_ack_rate'] = (packet_stats['duplicate_acks'] / packet_stats['tcp_packets']) * 100
    else:
        packet_stats['retransmission_rate'] = 0
        packet_stats['out_of_order_rate'] = 0
        packet_stats['duplicate_ack_rate'] = 0

    # Calcular estatísticas por intervalo de tempo
    intervals = sorted(packet_stats['time_intervals'].keys())
    packets_per_interval = [packet_stats['time_intervals'][i]['packets'] for i in intervals]
    retrans_per_interval = [packet_stats['time_intervals'][i]['retransmissions'] for i in intervals]
    bytes_per_interval = [packet_stats['time_intervals'][i]['bytes'] for i in intervals]

    if len(packets_per_interval) > 0:
        packet_stats['avg_packets_per_interval'] = statistics.mean(packets_per_interval)
        packet_stats['max_packets_per_interval'] = max(packets_per_interval)
        packet_stats['min_packets_per_interval'] = min(packets_per_interval)

        # Calcular taxas de retransmissão por intervalo
        retrans_rates = []
        for i in range(len(packets_per_interval)):
            if packets_per_interval[i] > 0:
                rate = (retrans_per_interval[i] / packets_per_interval[i]) * 100
            else:
                rate = 0
            retrans_rates.append(rate)

        packet_stats['avg_retrans_rate_per_interval'] = statistics.mean(retrans_rates)
        packet_stats['max_retrans_rate_per_interval'] = max(retrans_rates) if retrans_rates else 0
        packet_stats['min_retrans_rate_per_interval'] = min(retrans_rates) if retrans_rates else 0

        # Taxa de transferência (bytes/segundo)
        throughput = [b / interval_seconds for b in bytes_per_interval]
        packet_stats['avg_throughput'] = statistics.mean(throughput)
        packet_stats['max_throughput'] = max(throughput)
        packet_stats['min_throughput'] = min(throughput)

    # Calcular estatísticas por fluxo TCP
    for stream in packet_stats['tcp_streams'].values():
        if stream['packet_count'] > 0:
            stream['retrans_rate'] = (stream['retrans_count'] / stream['packet_count']) * 100
            stream['out_of_order_rate'] = (stream['out_of_order_count'] / stream['packet_count']) * 100
            stream['dup_ack_rate'] = (stream['dup_ack_count'] / stream['packet_count']) * 100

    return packet_stats


def generate_packet_loss_reports(packet_stats):
    """Gera relatórios e gráficos com base nas estatísticas de perda de pacotes"""
    # Criar DataFrames para análise
    df_summary = create_summary_dataframe(packet_stats)
    df_intervals = create_interval_dataframe(packet_stats)
    df_streams = create_stream_dataframe(packet_stats)

    # Gerar relatório Excel
    generate_excel_report(df_summary, df_intervals, df_streams)

    # Gerar gráficos
    plot_packet_loss_trends(df_intervals)
    plot_stream_analysis(df_streams)

    # Mostrar resumo no console
    print("\n=== RESUMO DE PERDA DE PACOTES ===")
    print(df_summary)
    print("\nTop 10 Fluxos TCP com Maior Taxa de Retransmissão:")
    print(df_streams.sort_values('Retrans Rate', ascending=False).head(10))


def create_summary_dataframe(packet_stats):
    """Cria um DataFrame com estatísticas resumidas"""
    data = {
        'Métrica': [
            'Total de Pacotes',
            'Pacotes TCP',
            'Pacotes UDP',
            'Pacotes ICMP',
            'Outros Pacotes',
            'Retransmissões TCP',
            'Pacotes Fora de Ordem',
            'ACKs Duplicados',
            'Taxa de Retransmissão (%)',
            'Taxa de Fora de Ordem (%)',
            'Taxa de ACKs Duplicados (%)',
            'Taxa Média de Transferência (bytes/s)'
        ],
        'Valor': [
            packet_stats['total_packets'],
            packet_stats['tcp_packets'],
            packet_stats['udp_packets'],
            packet_stats['icmp_packets'],
            packet_stats['other_packets'],
            packet_stats['retransmissions'],
            packet_stats['out_of_order'],
            packet_stats['duplicate_acks'],
            packet_stats.get('retransmission_rate', 0),
            packet_stats.get('out_of_order_rate', 0),
            packet_stats.get('duplicate_ack_rate', 0),
            packet_stats.get('avg_throughput', 0)
        ]
    }
    return pd.DataFrame(data)


def create_interval_dataframe(packet_stats):
    """Cria um DataFrame com dados por intervalo de tempo"""
    intervals = sorted(packet_stats['time_intervals'].keys())
    data = {
        'Intervalo': [],
        'Pacotes': [],
        'Retransmissões': [],
        'Fora de Ordem': [],
        'Bytes': [],
        'Taxa Retrans (%)': [],
        'Throughput (bytes/s)': []
    }

    for interval in intervals:
        stats = packet_stats['time_intervals'][interval]
        data['Intervalo'].append(interval)
        data['Pacotes'].append(stats['packets'])
        data['Retransmissões'].append(stats['retransmissions'])
        data['Fora de Ordem'].append(stats['out_of_order'])
        data['Bytes'].append(stats['bytes'])

        if stats['packets'] > 0:
            retrans_rate = (stats['retransmissions'] / stats['packets']) * 100
        else:
            retrans_rate = 0
        data['Taxa Retrans (%)'].append(retrans_rate)

        # Assumindo intervalo de 5 segundos (padrão da função)
        data['Throughput (bytes/s)'].append(stats['bytes'] / 5)

    return pd.DataFrame(data)


def create_stream_dataframe(packet_stats):
    """Cria um DataFrame com dados por fluxo TCP"""
    streams = []
    for stream_key, stats in packet_stats['tcp_streams'].items():
        streams.append({
            'Origem': f"{stats['src']}:{stats['src_port']}",
            'Destino': f"{stats['dst']}:{stats['dst_port']}",
            'Pacotes': stats['packet_count'],
            'Retransmissões': stats['retrans_count'],
            'Fora de Ordem': stats['out_of_order_count'],
            'ACKs Duplicados': stats['dup_ack_count'],
            'Retrans Rate (%)': stats.get('retrans_rate', 0),
            'Out of Order Rate (%)': stats.get('out_of_order_rate', 0),
            'Dup ACK Rate (%)': stats.get('dup_ack_rate', 0)
        })
    return pd.DataFrame(streams)


def generate_excel_report(df_summary, df_intervals, df_streams):
    """Gera um relatório Excel completo"""
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f'relatorio_perda_pacotes_{timestamp}.xlsx'

    wb = Workbook()

    # Planilha de resumo
    ws_summary = wb.active
    ws_summary.title = "Resumo"
    ws_summary.append(["Relatório de Perda de Pacotes"])
    ws_summary.append(["Data/Hora", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws_summary.append([])

    for r in dataframe_to_rows(df_summary, index=False, header=True):
        ws_summary.append(r)

    # Planilha de intervalos de tempo
    ws_intervals = wb.create_sheet("Análise Temporal")
    for r in dataframe_to_rows(df_intervals, index=False, header=True):
        ws_intervals.append(r)

    # Planilha de fluxos TCP
    ws_streams = wb.create_sheet("Fluxos TCP")
    for r in dataframe_to_rows(df_streams, index=False, header=True):
        ws_streams.append(r)

    # Formatação
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

    # Salvar
    wb.save(filename)
    print(f"\nRelatório Excel gerado: {filename}")


def plot_packet_loss_trends(df_intervals):
    """Gera gráficos de tendência temporal"""
    plt.figure(figsize=(14, 10))

    # Gráfico 1: Pacotes e Retransmissões ao longo do tempo
    plt.subplot(2, 1, 1)
    plt.plot(df_intervals['Intervalo'], df_intervals['Pacotes'], label='Pacotes', marker='o')
    plt.plot(df_intervals['Intervalo'], df_intervals['Retransmissões'], label='Retransmissões', marker='x')
    plt.title('Pacotes e Retransmissões por Intervalo de Tempo')
    plt.xlabel('Intervalo de Tempo')
    plt.ylabel('Contagem')
    plt.legend()
    plt.grid(True)

    # Gráfico 2: Taxa de retransmissão e throughput
    plt.subplot(2, 1, 2)
    plt.plot(df_intervals['Intervalo'], df_intervals['Taxa Retrans (%)'], label='Taxa de Retransmissão (%)',
             color='red', marker='o')
    plt.twinx()
    plt.plot(df_intervals['Intervalo'], df_intervals['Throughput (bytes/s)'], label='Throughput (bytes/s)',
             color='green', marker='x')
    plt.title('Taxa de Retransmissão e Throughput por Intervalo de Tempo')
    plt.xlabel('Intervalo de Tempo')
    plt.ylabel('Taxa/Throughput')
    plt.legend()
    plt.grid(True)

    plt.tight_layout()
    filename = 'tendencia_perda_pacotes.png'
    plt.savefig(filename, dpi=300, bbox_inches='tight')
    plt.show()
    print(f"Gráfico de tendências salvo como: {filename}")


def plot_stream_analysis(df_streams):
    """Gera gráficos de análise por fluxo TCP"""
    if len(df_streams) == 0:
        return

    # Pegar os top 20 fluxos com mais retransmissões
    top_streams = df_streams.nlargest(20, 'Retransmissões')

    plt.figure(figsize=(14, 8))

    # Gráfico de barras para taxas de retransmissão
    plt.barh(
        top_streams.apply(lambda x: f"{x['Origem']} → {x['Destino']}", axis=1),
        top_streams['Retrans Rate (%)'],
        color='salmon',
        edgecolor='black'
    )

    plt.title('Top 20 Fluxos TCP por Taxa de Retransmissão')
    plt.xlabel('Taxa de Retransmissão (%)')
    plt.ylabel('Fluxo TCP')
    plt.grid(axis='x', alpha=0.4)
    plt.tight_layout()

    filename = 'analise_fluxos_tcp.png'
    plt.savefig(filename, dpi=300, bbox_inches='tight')
    plt.show()
    print(f"Gráfico de fluxos TCP salvo como: {filename}")


# Exemplo de uso
if __name__ == "__main__":
    stats = analyze_packet_loss('captura.pcapng', analysis_interval=5)